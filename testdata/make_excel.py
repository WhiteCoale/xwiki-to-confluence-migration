#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate the input Excel files for the migration tool.

Anforderungen.txt: "Als Eingabe gibt es einen Ordner gefuellt mit Excel Dateien.
Je Excel Datei ist in Spalte A ein Status ob der xWiki Artikel 'Migrieren' soll
oder 'Loeschen'".

All users, e-mail addresses and host names below are fictional.

The column layout is:

    A Status | B Kommentarfeld | C title | D initial_creator | E created_date
    F last_editor | G last_change_date | H path | I link

Written to ./input/ as a formatted Excel table.

The paths correspond 1:1 to the pages created by seed_xwiki.py, except for the
deliberate edge cases in 03_Sonderfaelle.xlsx.

Usage:
    python testdata/make_excel.py
"""

import os
import urllib.parse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

XWIKI_BASE = "http://localhost:8080/bin/view"
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "input")

HEADERS = ["Status", "Kommentarfeld", "title", "initial_creator", "created_date",
           "last_editor", "last_change_date", "path", "link"]

AK = "anna.mueller@example.com"
HS = "thomas.berger@example.com"
FW = "petra.klein@example.com"
MB = "jonas.wagner@example.com"
LEGACY = "XWiki.legacy001"


def link_for(path):
    """Build the local xWiki view URL from the page path."""
    segments = path.split("/")
    if segments[-1] == "WebHome":
        segments = segments[:-1]
        trailing = "/"
    else:
        trailing = ""
    return "{}/{}{}".format(
        XWIKI_BASE, "/".join(urllib.parse.quote(s) for s in segments), trailing)


def row(status, kommentar, title, creator, created, editor, changed, path):
    return [status, kommentar, title, creator, created, editor, changed, path,
            link_for(path)]


# --------------------------------------------------------------------------
# 01 - Clientthemen
# --------------------------------------------------------------------------
CLIENTTHEMEN = [
    row("Migrieren", "Ja, aktualisierung notwendig",
        "PTS Protokolltestsystem auf Techniker Client", AK,
        "2025-11-19T09:35:31Z", AK, "2025-11-19T09:35:31Z",
        "Clientthemen/PTS Protokolltestsystem auf Techniker Client/WebHome"),
    row("Migrieren", "Ja, aktualisierung notwendig",
        "Hyper-V Host unter Accenture Clients", AK,
        "2025-07-01T15:02:08Z", AK, "2025-07-01T15:02:50Z",
        "Clientthemen/Hyper-V Host unter Accenture/WebHome"),
    row("Migrieren", "Ja", "Mainboard Tausch - Rechnerwiederinbetriebnahme", AK,
        "2025-03-05T13:52:56Z", AK, "2025-03-26T13:12:20Z",
        "Clientthemen/Mainboard Tausch/WebHome"),
    row("Migrieren", "Ja, Ueberordner soll als Seite mit wandern", "Clientthemen", AK,
        "2018-11-02T08:00:00Z", MB, "2025-12-02T10:15:00Z",
        "Clientthemen/WebHome"),
    row("Migrieren", "Ja, Unterordner", "Windows", AK,
        "2019-01-15T07:45:00Z", AK, "2024-08-19T13:20:00Z",
        "Clientthemen/Windows/WebHome"),
    row(u"Löschen", "Nein", "aktuelle Windows 10 Enterprise Iso downloaden", LEGACY,
        "2019-05-06T11:24:20Z", AK, "2022-04-21T09:13:50Z",
        "Clientthemen/Windows/aktuelle Windows 10 Enterprise Iso downloaden/WebHome"),
    row("Migrieren", "Ja, Unterordner", "VMWare Client", AK,
        "2020-02-11T09:10:00Z", AK, "2023-05-04T16:02:00Z",
        "Clientthemen/VMWare Client/WebHome"),
    row("Migrieren", "Ja, Vmware ist in Projekten noch aktuell",
        "VMWare Side Channel Mitigation", AK,
        "2022-01-27T12:01:07Z", AK, "2022-01-27T12:01:07Z",
        "Clientthemen/VMWare Client/VMWare Side Channel Mitigation/WebHome"),
    row(u"Löschen", "Nein, eher nicht von mir sondern Micha",
        "Taschenrechner-App reparieren Windows 10", FW,
        "2019-10-01T07:45:18Z", AK, "2021-07-15T12:01:53Z",
        "Clientthemen/Taschenrechner-App reparieren Windows 10/WebHome"),
    row(u"Löschen", "Nein", "Energieoptionen beim Betrieb mit Stromversorgung", AK,
        "2021-06-15T06:16:23Z", AK, "2021-06-15T06:52:55Z",
        "Clientthemen/Energieoptionen beim Betrieb mit Stromversorgung/WebHome"),
    row("Migrieren", "Ja, Unterordner", "Drucker", MB,
        "2021-09-08T11:11:00Z", MB, "2024-03-17T09:05:00Z",
        "Clientthemen/Drucker/WebHome"),
    # Duplikat-Titel 1/2 - gleicher Titel wie Serverthemen/Backup/Installation
    row("Migrieren", "Ja - Achtung: Titel 'Installation' existiert mehrfach",
        "Installation", MB, "2021-09-08T11:30:00Z", MB, "2024-03-17T09:12:00Z",
        "Clientthemen/Drucker/Installation"),
]

# --------------------------------------------------------------------------
# 02 - Serverthemen
# --------------------------------------------------------------------------
SERVERTHEMEN = [
    row("Migrieren", "Ja, dann wissens alle :D", "Toolbox Installationsanleitung", AK,
        "2023-06-27T08:30:37Z", AK, "2023-06-27T08:32:41Z",
        "Serverthemen/Toolbox Installationsanleitung/WebHome"),
    row("Migrieren", "Ja eher von Hans", "Server verschwindet aus WSUS", HS,
        "2019-01-30T13:06:18Z", AK, "2021-04-22T09:12:42Z",
        "Serverthemen/Wsus/Server verschwindet aus WSUS"),
    row("Migrieren", "Ja eher von Hans", "Patchen der Windows Server mittels Wsus", HS,
        "2019-01-30T13:06:24Z", AK, "2021-03-02T08:34:10Z",
        "Serverthemen/Wsus/Patchen der Windows Server mittels Wsus/WebHome"),
    row("Migrieren", "Ja, Ueberordner", "Serverthemen", HS,
        "2018-10-05T06:30:00Z", AK, "2025-10-14T08:44:00Z",
        "Serverthemen/WebHome"),
    row("Migrieren", "Ja, Unterordner", "Wsus", HS,
        "2019-01-30T13:00:00Z", HS, "2022-11-08T15:31:00Z",
        "Serverthemen/Wsus/WebHome"),
    row("Migrieren", "Ja, Unterordner", "Backup", HS,
        "2020-04-21T10:00:00Z", HS, "2025-02-19T11:47:00Z",
        "Serverthemen/Backup/WebHome"),
    # Duplikat-Titel 2/2
    row("Migrieren", "Ja - zweiter Treffer fuer den Titel 'Installation'",
        "Installation", HS, "2020-04-21T10:20:00Z", HS, "2025-02-19T11:52:00Z",
        "Serverthemen/Backup/Installation"),
]

# --------------------------------------------------------------------------
# 03 - Sonderfaelle / Datenqualitaet
# --------------------------------------------------------------------------
SONDERFAELLE = [
    # Versteckte Benutzerseite - muss laut Anforderung migriert werden
    row("Migrieren", "Versteckte Seite, soll trotzdem mit", u"Persönliche Notizen (versteckt)",
        FW, "2024-06-03T14:20:00Z", FW, "2025-01-09T08:12:00Z",
        "PersoenlicheNotizen/VersteckteSeite"),
    row("Migrieren", "Benutzerbereich (versteckt)", u"Persönliche Notizen", FW,
        "2024-06-03T14:18:00Z", FW, "2024-06-03T14:18:00Z",
        "PersoenlicheNotizen/WebHome"),
    # Steht in der Liste, existiert aber nicht mehr in xWiki
    row("Migrieren", "Seite wurde zwischenzeitlich in xWiki geloescht",
        "Citrix Receiver Altinstallation", AK,
        "2017-03-14T09:00:00Z", AK, "2018-07-30T12:00:00Z",
        "Clientthemen/Citrix Receiver Altinstallation/WebHome"),
    # Doppelter Eintrag zur selben Seite (kommt schon in 01 vor)
    row("Migrieren", "Doppelt erfasst - Dublette zu 01_Clientthemen.xlsx",
        "Mainboard Tausch - Rechnerwiederinbetriebnahme", AK,
        "2025-03-05T13:52:56Z", AK, "2025-03-26T13:12:20Z",
        "Clientthemen/Mainboard Tausch/WebHome"),
    # Widerspruechlicher Status zur selben Seite wie oben
    row(u"Löschen", "Widerspruch: in 01 als 'Migrieren' gefuehrt",
        "Mainboard Tausch - Rechnerwiederinbetriebnahme", AK,
        "2025-03-05T13:52:56Z", AK, "2025-03-26T13:12:20Z",
        "Clientthemen/Mainboard Tausch/WebHome"),
    # Kleinschreibung / Leerzeichen im Status
    row(u"  migrieren ", "Status klein geschrieben und mit Leerzeichen",
        u"Monitoring Übersicht", MB, "2023-08-01T09:00:00Z", MB,
        "2025-06-11T07:25:00Z", "Serverthemen/Monitoring/WebHome"),
    # Leerer Status
    row("", "Status nicht gepflegt", u"Notebook-Übergabe & -Rückgabe (Größe/Prüfung)",
        MB, "2024-02-05T10:00:00Z", AK, "2025-04-30T16:40:00Z",
        "Clientthemen/Notebook Uebergabe/WebHome"),
    # Unbekannter Status
    row("unklar", "Muss noch geklaert werden", "Drucker", MB,
        "2021-09-08T11:11:00Z", MB, "2024-03-17T09:05:00Z",
        "Clientthemen/Drucker/WebHome"),
    # Muellzeile wie in der Beispieldatei (dort Zeile 5)
    ["Migrieren"] * 9,
    # xWiki Systemseite - darf nicht migriert werden
    row("Migrieren", "Systemseite, faelschlich in der Liste", "XWiki Preferences",
        LEGACY, "2015-01-01T00:00:00Z", LEGACY, "2015-01-01T00:00:00Z",
        "XWiki/XWikiPreferences"),
]


def write_workbook(filename, rows, table_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "xWiki Artikel"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = [12, 42, 46, 32, 22, 32, 22, 62, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ref = "A1:{}{}".format(get_column_letter(len(HEADERS)), len(rows) + 1)
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A2"

    path = os.path.join(OUT_DIR, filename)
    wb.save(path)
    print("  {}  ({} Zeilen)".format(path, len(rows)))


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    print("Schreibe Eingabe-Excel-Dateien nach {} ...".format(OUT_DIR))
    write_workbook("01_Clientthemen.xlsx", CLIENTTHEMEN, "XWikiPagesClient")
    write_workbook("02_Serverthemen.xlsx", SERVERTHEMEN, "XWikiPagesServer")
    write_workbook("03_Sonderfaelle.xlsx", SONDERFAELLE, "XWikiPagesSonder")
    total = len(CLIENTTHEMEN) + len(SERVERTHEMEN) + len(SONDERFAELLE)
    print("Fertig: 3 Dateien, {} Datenzeilen.".format(total))


if __name__ == "__main__":
    main()
